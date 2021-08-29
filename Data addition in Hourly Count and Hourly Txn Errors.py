import time,os,config,sys
from openpyxl.styles import Alignment,Font,Border,Side
#import importlib
#importlib.reload(config)

import mysql.connector
mydb= mysql.connector.connect(host=config.HOST, user=config.USER, password=config.PASSWORD, database=config.DATABASE)
cur= mydb.cursor()
cur2= mydb.cursor()

os.chdir(config.PATH)
print(os.getcwd())
Files=os.listdir()

date1=input('enter the date in dd-mm-yyyy format:' )
time1=input('enter the time1 in hh:mm format:' )
time2=input('enter the time2 in hh:mm format:' )

start=time.time()		#start time

d1= date1.split("-")
t1= time1.split(":")
t2= time2.split(":")

dict1={'23':5, '00':6, '01':7, '02':8, '03':9, '04':10, '05':11, '06':12, '07':13, '08':14, '09':15, '10':16, '11':17, '12':18, '13':19, '14':20, '15':21, '16':22,
      '17':23, '18':24, '19':25, '20':26, '21':27, '22':28}   #to find columns in worksheet as per time/hour provided

dict2={'00':'12 AM', '01':'01 AM', '02':'02 AM', '03':'03 AM', '04':'04 AM', '05':'05 AM', '06':'06 AM', '07':'07 AM', '08':'08 AM', '09':'09 AM', '10':'10 AM', '11':'11 AM', '12':'12 PM', '13':'01 PM', '14':'02 PM', '15':'03 PM', '16':'04 PM',
      '17':'05 PM', '18':'06 PM', '19':'07 PM', '20':'08 PM', '21':'09 PM', '22':'10 PM','23':'11 PM', '24':'12 AM', } '''dict2 is replacing 24hr format with 12hr format'''

def nextDayDate(date1):
    '''function to find next day date '''
	
    d1=date1.split("-")   
    if d1[1]=='12' and d1[0]=='31':
        d1[0]='01'
        d1[1]='01'
        d1[2]=str(int(d1[2])+1)
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1
        #break----remember this to add
    
    elif d1[0]>='28' and d1[0]<='29' and d1[1]=='02':
        d1[0]='01'
        d1[1]='03'
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1
    
    elif d1[0]>='30' and d1[1]=='04' or d1[0]>='30'and d1[1]=='06' or d1[0]>='30'and d1[1]=='09' or d1[0]>='30'and d1[1]=='11':
        d1[0]='01'
        if d1[1]=='11' or d1[1]=='09':
            d1[1]=str(int(d1[1])+1)   #increment, otherwise it will result in '010' instead of '10'
        else:
            d1[1]='0'+str(int(d1[1])+1)            
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1

    elif (d1[0]>='31' and d1[1]=='01'or d1[0]>='31' and d1[1]=='03' or d1[0]>='31' and d1[1]=='05' or d1[0]>='31' and d1[1]=='07' or d1[0]>='31' and d1[1]=='08'
         or d1[0]>='31' and d1[1]=='10' or d1[0]>='31' and d1[1]=='12'):
        d1[0]='01'
        if d1[1]>='01' and d1[1]<'09':
            d1[1]='0'+str(int(d1[1])+1)   #increment, otherwise it will result in '010' instead of '10'
        else:
            d1[1]=str(int(d1[1])+1)
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1    
    
    elif d1[0]>='01' and d1[0]<'09':
        d1[0]='0'+str(int(d1[0])+1) #increment, otherwise it will result in '010' instead of '10'
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2] #newdate
        #print(dt1)
        return dt1
    else:
        d1[0]=str(int(d1[0])+1) #increment
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2] #newdate
        #print(dt1)
        return dt1

for file in Files:
    if time1=='23:00' and time2=='24:00':
        if file[26:36]==nextDayDate(date1):
            filename= file      
            
    elif file[26:36]==date1:
            filename= file
try:
    import openpyxl as op
    excel= op.load_workbook(filename)  	#loading existing excel file
    ws=excel['Hourly_Count']		#worksheet 1
    ws1=excel["Hourly_Txn_Erros"]	#worksheet 2

    while True:
        if t1[0]==t2[0]:
            break

        else:
            #TRANSACTION COUNT worksheet
            a=5				#row number
            b=dict1[t1[0]]		#column number

            s='''select count(*),sum(Amount)
            from my_table where original_transaction_time between
            str_to_date('%s-%s-%s %s:00:00', '%d-%m-%Y %H:%i:%S') and 
            str_to_date('%s-%s-%s %s:59:59', '%d-%m-%Y %H:%i:%S') and txn_type= 'P2A_INWARD' '''

            cur.execute(s,(int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0]),int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0])))
            result= cur.fetchall()

            if result[0][0]==0 and result[0][1]==None:
                for r in range(2):
                    ws.cell(a,b).value=0
                    a+=1
            else:
                for tuple1 in result:
                    for value in tuple1:
                        ws.cell(a,b,value)
                        a+=1

            s1= ''' select count(*),sum(Amount)
            from my_table where original_transaction_time between
            str_to_date('%s-%s-%s %s:00:00', '%d-%m-%Y %H:%i:%S') and 
            str_to_date('%s-%s-%s %s:59:59', '%d-%m-%Y %H:%i:%S') and txn_type= 'P2A_INWARD' and status='A' '''        

            cur.execute(s1,(int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0]),int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0])))
            result= cur.fetchall()

            if result[0][0]==0 and result[0][1]==None:
                for r in range(2):
                    ws.cell(a,b).value=0
                    a+=1
            else:
                for tuple1 in result:
                    for value in tuple1:
                        ws.cell(a,b,value)
                        a+=1
                    
            a=a+6				#To move to outward cells
            for cha in config.CHANNELS:         #cha=channel

                s2= ''' select count(*),sum(Amount)
                from my_table where original_transaction_time between
                str_to_date('%s-%s-%s %s:00:00', '%d-%m-%Y %H:%i:%S') and 
                str_to_date('%s-%s-%s %s:59:59', '%d-%m-%Y %H:%i:%S') and txn_type= 'P2A_OUTWARD' and channel= %s '''

                cur.execute(s2,(int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0]),int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0]),cha))
                result= cur.fetchall()

                if result[0][0]==0 and result[0][1]==None:
                    for r in range(2):
                        ws.cell(a,b).value=0
                        a+=1
                else:
                    for tuple1 in result:
                        for value in tuple1:
                            ws.cell(a,b,value)
                            a+=1
	    
	    #fetching approved transactions data
	
            s3=''' select count(*),sum(Amount)
            from my_table where original_transaction_time between
            str_to_date('%s-%s-%s %s:00:00', '%d-%m-%Y %H:%i:%S') and 
            str_to_date('%s-%s-%s %s:59:59', '%d-%m-%Y %H:%i:%S') and txn_type= 'P2A_OUTWARD' and status='A' '''

            cur.execute(s3,(int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0]),int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0])))
            result= cur.fetchall()

            if result[0][0]==0 and result[0][1]==None:
                for r in range(2):
                    ws.cell(a,b).value=0
                    a+=1
            else:
                for tuple1 in result:
                    for value in tuple1:
                        ws.cell(a,b,value)
                        a+=1

            #TODAY'S ERROR worksheet

            dim=ws1.calculate_dimension()
            dim1=dim.split(":")
            
            if t1[0]>='00' and t1[0]<'09':
                t12='0'+str(int(t1[0])+1)  #increment, otherwise it will result in '010' instead of '10'
            else:
                t12=str(int(t1[0])+1)

            v1= dict2[t1[0]]+' - '+dict2[t12] 

            x=int(dim1[1][1:])		#row number
            y=3				#column number

            ws1.cell(x+1,2,'Inward')
            ws1.cell(x+1,1,v1)
            
            s4='''select error_message, error_id, count(*)
            from my_table where original_transaction_time between
            str_to_date('%s-%s-%s %s:00:00', '%d-%m-%Y %H:%i:%S') and 
            str_to_date('%s-%s-%s %s:59:59', '%d-%m-%Y %H:%i:%S') and 
            status != 'A' and  txn_type='P2A_INWARD' group by error_message,error_id''' #fetching inward errors

            cur2.execute(s4,(int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0]),int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0])))
            result=cur2.fetchall()

            if len(result)==0:
                x+=1
                for tuple1 in range(3):                            
                    ws1.cell(x,y).value='nil'
                    y+=1    
                y=3
                count1=1
            else:    
                count1=0
                for tuple1 in result:
                    x+=1
                    for value in tuple1:        
                        ws1.cell(x,y,value)
                        y+=1    
                    y=3
                    count1 +=1
                
            ws1.merge_cells(start_row=(x-count1+1), start_column=2, end_row=x, end_column=2)
            ws1.cell(x+1,2,'Outward')

            s5='''select error_message, error_id, count(*)
            from my_table where original_transaction_time between
            str_to_date('%s-%s-%s %s:00:00', '%d-%m-%Y %H:%i:%S') and 
            str_to_date('%s-%s-%s %s:59:59', '%d-%m-%Y %H:%i:%S') and 
            status != 'A' and  txn_type='P2A_OUTWARDUTWARDUTWARDUTWARD' group by error_message,error_id ''' #fetching outward errors

            cur2.execute(s5,(int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0]),int(d1[0]),int(d1[1]),int(d1[2]),int(t1[0])))
            result=cur2.fetchall()

            if len(result)==0:
                x+=1
                for tuple1 in range(3):                            
                    ws1.cell(x,y).value='nil'
                    y+=1    
                y=3
                count2=1
                
            else:
                count2=0
                for tuple1 in result:
                    x+=1
                    for value in tuple1:        
                        ws1.cell(x,y,value)
                        y+=1    
                    y=3
                    count2 +=1
            ws1.merge_cells(start_row=(x-count2+1), start_column=2, end_row=x, end_column=2)
            ws1.merge_cells(start_row=(x-count1-count2+1), start_column=1, end_row=x, end_column=1)
            excel.save(filename)    
            
	    #Changing the time/hour			                
            if t1[0]>='00' and t1[0]<'09':            
                t1[0]='0'+ str(int(t1[0])+1)
            else:
                t1[0]=str(int(t1[0])+1)
            a=5  	#reseting the row number  
        
    for col in ws1.columns:	#doing basic formatting like center align, border etc.
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                thin = Side(border_style="thin", color="000000")
                #double = Side(border_style="double", color="000000")
                cell.border=Border(top=thin, left=thin, right=thin, bottom=thin)

    excel.save(filename)

#FINDING EXCEPTION AND PASSING APPROPRIATE MESSAGE
except NameError:
    if time1=='23:00' and time2=='24:00':
        print('\nExcel file for date',nextDayDate(date1),'not found in current working directory \nChange the file location PATH in config if needed.')
    else:
        print('\nExcel file for date',date1,'not found in current working directory \nChange the file location PATH in config if needed.')
          
except PermissionError:
    if time1=='23:00' and time2=='24:00':        
        print('\nYour excel file for date',nextDayDate(date1),'is OPEN, please close it and try again!')
    else:
        print('\nYour excel file for date',date1,'is OPEN, please close it and try again!')

except:
    print('Oops! ',sys.exc_info(),'occured.')

finally:    
    cur.close()
    mydb.close()

end=time.time()
print('execution time is:',(end-start))
        
