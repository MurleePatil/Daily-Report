import time,os,datetime,sys
import config

import mysql.connector
mydb= mysql.connector.connect(host=config.HOST, user=config.USER, password=config.PASSWORD, database=config.DATABASE)
cur= mydb.cursor()
cur2= mydb.cursor()

os.chdir(config.PATH)
print(os.getcwd())
Files=os.listdir()

date1=input('enter the date in dd-mm-yyyy format:' )
d1= date1.split("-")

start=time.time()

def nextDayDate(date1):
    ''' function to find next day date '''
    d1=date1.split("-")
    if d1[1]=='12' and d1[0]=='31':
        d1[0]='01'
        d1[1]='01'
        d1[2]=str(int(d1[2])+1)
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1
    
    elif d1[0]>='28' and d1[0]<='29' and d1[1]=='02':
        d1[0]='01'
        d1[1]='03'
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1
    
    elif d1[0]>='30' and d1[1]=='04' or d1[0]>='30'and d1[1]=='06' or d1[0]>='30'and d1[1]=='09' or d1[0]>='30'and d1[1]=='11':
        d1[0]='01'
        if d1[1]=='11' or d1[1]=='09':
            d1[1]=str(int(d1[1])+1)	#increment, otherwise it will result in '010' instead of '10'
        else:
            d1[1]='0'+str(int(d1[1])+1)            
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1

    elif (d1[0]>='31' and d1[1]=='01'or d1[0]>='31' and d1[1]=='03' or d1[0]>='31' and d1[1]=='05' or d1[0]>='31' and d1[1]=='07' or d1[0]>='31' and d1[1]=='08'
         or d1[0]>='31' and d1[1]=='10' or d1[0]>='31' and d1[1]=='12'):
        d1[0]='01'
        if d1[1]>='01' and d1[1]<'09':
            d1[1]='0'+str(int(d1[1])+1)	#increment, otherwise it will result in '010' instead of '10'
        else:
            d1[1]=str(int(d1[1])+1)
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1    
    
    elif d1[0]>='01' and d1[0]<='09':
        d1[0]='0'+str(int(d1[0])+1) 	#increment, otherwise it will result in '010' instead of '10'
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2] 	#newdate
        #print(dt1)
        return dt1
    else:
        d1[0]=str(int(d1[0])+1) 	#increment, otherwise it will result in '010' instead of '10'
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2] 	#newdate
        #print(dt1)
        return dt1


for file in Files:    
    if file[26:36]==nextDayDate(date1):
        filename= file
        
class RowNotFoundError(Exception):
    pass

try:        
    import openpyxl as op
    excel= op.load_workbook(filename)	#loading the excel file	
    ws=excel['EOD_Count']		#worksheet

    
    def getRowNumber1(date1):
        i=0
        for row in ws.iter_rows(1,ws.max_row,1,1,True):
            i+=1
            j=2
            for cell in row:
                if type(cell) == datetime.datetime:       
                    if cell.strftime("%d-%m-%Y")==date1:
                        return i
    if getRowNumber1(date1)==None:
        raise RowNotFoundError
    else:
        print('Row number(EOD):',getRowNumber1(date1))

    a=getRowNumber1(date1)		#row number
    b=2					#cloumn number
    for tt in ['P2A_INWARD','P2A_OUTWARD']:    	#tt=transaction type
        
	#QUERY TO FETCH TOTAL DATA
        s='''select count(*), sum(Amount)
        from my_table where original_transaction_time between
        str_to_date('%s-%s-%s 00:00:00', '%d-%m-%Y %H:%i:%S') and 
        str_to_date('%s-%s-%s 23:59:59', '%d-%m-%Y %H:%i:%S') and txn_type=%s '''
	
	#EXECUTING THE QUERY AND FETCHING RESULTS
        cur.execute(s,(int(d1[0]),int(d1[1]),int(d1[2]),int(d1[0]),int(d1[1]),int(d1[2]),tt))
        result= cur.fetchall()

	#FILLING THE DATA IN EXCEL FILE        
        if len(result)==0:
            for r in range(2):
                ws.cell(a,b).value=0
                b+=1
        else:
            for tuple1 in result:
                for value in tuple1:
                    ws.cell(a,b,value)
                    b+=1

	#QUERY TO FETCH APPROVED DATA
        s1='''select count(*), sum(Amount)
        from my_table where original_transaction_time between
        str_to_date('%s-%s-%s 00:00:00', '%d-%m-%Y %H:%i:%S') and 
        str_to_date('%s-%s-%s 23:59:59', '%d-%m-%Y %H:%i:%S') and txn_type=%s and status='A' '''

        cur.execute(s1,(int(d1[0]),int(d1[1]),int(d1[2]),int(d1[0]),int(d1[1]),int(d1[2]),tt))
        result= cur.fetchall()

        if len(result)==0:
            for r in range(2):
                ws.cell(a,b).value=0
                b+=1
        else:
            for tuple1 in result:
                for value in tuple1:
                    ws.cell(a,b,value)
                    b+=1

	#QUERY TO FETCH UNSUCCESSFUL TRANSACTIONS DATA
        s2='''select count(*), sum(Amount)
        from my_table where original_transaction_time between
        str_to_date('%s-%s-%s 00:00:00', '%d-%m-%Y %H:%i:%S') and 
        str_to_date('%s-%s-%s 23:59:59', '%d-%m-%Y %H:%i:%S') and txn_type=%s and status!='A' '''

        cur.execute(s2,(int(d1[0]),int(d1[1]),int(d1[2]),int(d1[0]),int(d1[1]),int(d1[2]),tt))
        result= cur.fetchall()

        if len(result)==0:
            for r in range(2):
                ws.cell(a,b).value=0
                b+=1
        else:
            for tuple1 in result:
                for value in tuple1:
                    ws.cell(a,b,value)
                    b+=1

    
    excel.save(filename)

#THROWING SPECIFIC EXCEPTIONS AND APPROPRIATE MESSAGES
except NameError:
    print('\nExcel file for date',nextDayDate(date1),'not found in current working directory \nChange the file location PATH in config if needed OR create a new file.')

          
except PermissionError:
    print('\nYour excel file for date',nextDayDate(date1),'is OPEN, please close it and try again!')

except RowNotFoundError:
    print('\nYour excel file for date',nextDayDate(date1),'is found,BUT row with date',date1,'NOT found in the worksheets(EOD/NRE)')

except:
    print('Oops! ',sys.exc_info(),'occured.')

finally:    
    cur.close()
    mydb.close()

end=time.time()
print('\nexecution time is:',(end-start))
