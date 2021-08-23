import time,os,config,sys
#import importlib
#importlib.reload(config)
start=time.time()

os.chdir(config.PATH1)
print(os.getcwd())
Files=os.listdir()

from datetime import datetime

#import pytz
#tz=pytz.timezone('Asia/Kolkata')     	#timezone
#d=datetime.now(tz)
##d=datetime.now()
##date1=d.strftime('%d-%m-%Y')
#date1='01-11-2019'

date1=input('Enter a date in (dd-mm-yyyy) format: ')

def nextDayDate(date1):
    d1=date1.split("-")   
    if d1[1]=='12' and d1[0]=='31':
        d1[0]='01'
        d1[1]='01'
        d1[2]=str(int(d1[2])+1)
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1
    
    elif d1[0]>='28' and d1[0]<='29' and d1[1]=='02':
        d1[0]='01'
        d1[1]='03'
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1
    
    elif d1[0]>='30' and d1[1]=='04' or d1[0]>='30'and d1[1]=='06' or d1[0]>='30'and d1[1]=='09' or d1[0]>='30'and d1[1]=='11':
        d1[0]='01'
        if d1[1]=='11' or d1[1]=='09':
            d1[1]=str(int(d1[1])+1)
        else:
            d1[1]='0'+str(int(d1[1])+1)            
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1

    elif (d1[0]>='31' and d1[1]=='01'or d1[0]>='31' and d1[1]=='03' or d1[0]>='31' and d1[1]=='05' or d1[0]>='31' and d1[1]=='07' or d1[0]>='31' and d1[1]=='08'
         or d1[0]>='31' and d1[1]=='10' or d1[0]>='31' and d1[1]=='12'):
        d1[0]='01'
        if d1[1]>='01' and d1[1]<'09':
            d1[1]='0'+str(int(d1[1])+1)
        else:
            d1[1]=str(int(d1[1])+1)
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2]
        return dt1    
    
    elif d1[0]>='01' and d1[0]<='09':
        d1[0]='0'+str(int(d1[0])+1) #increment, otherwise it will result in '010' instead of '10'
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2] #newdate
        #print(dt1)
        return dt1
    else:
        d1[0]=str(int(d1[0])+1) #increment
        dt1=d1[0]+'-'+d1[1]+'-'+d1[2] #newdate
        #print(dt1)
        return dt1
try:
    from shutil import copyfile

    source_file='IMPS_Daily_Status_Report_('+date1+').xlsx'
    destination_file='IMPS_Daily_Status_Report_('+nextDayDate(date1)+').xlsx'

    copyfile(source_file, destination_file)

    import openpyxl as op
    excel= op.load_workbook(destination_file)
    ws=excel['Transaction_Count']		#worksheet 1
    ws1=excel["Today's_Error"]			#worksheet 2

    for row in range(5,9):
        for col in range(5,29):
            ws.cell(row,col).value=None

    for row in range(15,25):
        for col in range(5,29):
            ws.cell(row,col).value=None
    ws.cell(3,5).value= date1+' 23:00:00 - '+nextDayDate(date1)+' 23:00:00'
    excel.save(destination_file)
          
    List1=ws1.merged_cells.ranges
    while True:    
        for i in List1:
            ws1.unmerge_cells(range_string=str(i))

        if len(List1)==0:
            break
    ws1.delete_rows(2,ws1.max_row)    
    excel.save(destination_file)
except:
    print('oops!',sys.exc_info(),'occured')
    

end=time.time()
print('execution time is:',(end-start))
